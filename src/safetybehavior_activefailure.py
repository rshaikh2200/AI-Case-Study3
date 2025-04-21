from openai import OpenAI
import os
import re
import PyPDF2
import openpyxl
import traceback

# You can still provide your API key however you'd like:
# (either hard-coded or via environment variables).
# For demonstration, we'll leave it as is:
API_KEY = "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"

# Initialize the OpenAI client using the new structure.
client = OpenAI(api_key=API_KEY)

# --------------- Configuration ---------------

# Directory containing the PDFs
PDF_DIRECTORY = "case studies"

# Path to the Excel file (existing) to append
EXCEL_FILE_PATH = "case_studies_part1_100.xlsx"

# We now detect any line beginning with these prefixes (case-insensitive)
CASE_HEADING_PREFIXES = ["the case", "case & commentary"]
# MODIFICATION: Add "case study" and "clinical sequence" so we can capture those headings too.
CASE_HEADING_PREFIXES.extend(["case study", "clinical sequence"])

SAFETY_BEHAVIORS_DEFINITIONS = """
Here is a list of possible safety behaviors and their definitions (select only one if applicable):
a. Colleague Feedback
   Definition: Ask your colleagues to review your work and offer assistance in reviewing the work of others.
b. Team Evaluation
   Definition: Reflect on what went well with the team, what didn't, how to improve, and who will follow through.
   All team members should freely speak up. A debrief typically lasts only 3 minutes.
c. Risk Intervention
   Definition: Ask a question to gently prompt the other person of a potential safety issue,
   request a change to make sure the person is fully aware of the risk.
   Voice a concern if the person is resistant. Use the chain of command if the possibility of patient harm persists.
d. Validation Assessment
   Definition: An internal Check (Does this make sense to me? Is it right, based on what I know?
   Is this what I expected? Does this information fit with my past experience or other information?).
   Then Verify (check with an independent qualified source).
e. SAFE (Stop-Assess-Focus-Engage)
   Definition: Stop (pause for 2 seconds to focus on task at hand),
   Assess (consider the action you're about to take),
   Focus (concentrate and carry out the task),
   Engage (check to make sure the task was done correctly and you got the right result).
f. Interruption Free Zone
   Definition:
   1) Avoid interrupting others while they are performing critical tasks
   2) Avoid distractions while completing critical tasks: Use phrases like 'Stand by' or 'Hold on'.
g. Effective Care Transitions
   Definition: Six important principles that make Effective Care Transitions:
   1) Standardized and streamlined
   2) Distraction-free environment
   3) Face-to-face/bedside (interactive)
   4) Acknowledgments/repeat backs
   5) Verbal with written/printed information
   6) Opportunity for questions/clarification

h. CARE (Communicate-Acknowledge-Repeat-Evaluate)
   Definition:
   1) Sender communicates information to receiver,
   2) Receiver listens or writes down the information and reads/repeats it back
      as written or heard to the sender,
   3) Sender then acknowledges the accuracy of the read-back by stating 'that's correct'.
      If not correct, the sender repeats/clarifies the communication again.
i. Clear Communications
   Definition: Consists of using clear letters and numbers in communication
   (e.g., replace 'fifteen' with 'one-five', use phonetic alphabet for letters).
j. Clarifying Information
   Definition: Requesting additional information and expressing concerns to avoid misunderstandings.
"""

PRIMARY_ACTIVE_FAILURE_DEFINITION = """
Definition of Primary Active Failure:
A primary active failure refers to an error that occurs directly at the point of care,
involving healthcare providers. These errors are immediate, observable,
and often have direct consequences for patient safety.
"""

def summarize_case_study(full_case_text: str) -> str:
    """
    Summarize the given case study in exactly 200 words using ChatGPT.
    """
    prompt = (
        "Please summarize the following case study in exactly 200 words:\n\n"
        f"{full_case_text}\n\n"
        "Your response should be 200 words, no more, no less."
    )
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role":"system","content":"You are a helpful assistant."},
                {"role":"user",  "content":prompt}
            ],
            max_tokens=4000,
            temperature=0.7
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        traceback.print_exc()
        return f"Summarization Error: {e}"

def identify_primary_active_failure_and_safety_behavior(full_case_text: str) -> (str, str):
    """
    Ask ChatGPT to:
      1) Identify the primary active failure.
      2) Choose exactly ONE 'Main Safety Behavior' from the list by letter and name.
    Uses multiple fallbacks to parse reliably.
    """
    prompt_text = (
        f"{PRIMARY_ACTIVE_FAILURE_DEFINITION}\n\n"
        f"{SAFETY_BEHAVIORS_DEFINITIONS}\n\n"
        "Based on the following case study, please:\n"
        "1) Identify the primary active failure.\n"
        "2) Choose exactly ONE 'Main Safety Behavior' from the list above by its letter and name (e.g., 'c. Risk Intervention').\n\n"
        "Format exactly as:\n"
        "Primary Active Failure: <short text>\n"
        "Main Safety Behavior: <letter>. <Behavior Name>\n\n"
        f"Case Study:\n{full_case_text}\n\n"
    )

    try:
        completion = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user",   "content": prompt_text}
            ],
            max_tokens=4000,
            temperature=0.7
        )
        content = completion.choices[0].message.content.strip()

        primary_active_failure = ""
        main_safety_behavior  = ""
        for line in content.splitlines():
            low = line.lower()
            if low.startswith("primary active failure:"):
                primary_active_failure = line.split(":",1)[1].strip()
            if low.startswith("main safety behavior:"):
                main_safety_behavior = line.split(":",1)[1].strip()

        # Fallback #1: lettered behaviors a-j
        if not main_safety_behavior:
            m = re.search(r"(?mi)^[a-j]\.\s*([A-Za-z ]+)$", content, re.MULTILINE)
            if m:
                main_safety_behavior = m.group(0).strip()

        # Fallback #2: scan known names
        if not main_safety_behavior:
            behaviors = [
                "Colleague Feedback",
                "Team Evaluation",
                "Risk Intervention",
                "Validation Assessment",
                "SAFE",
                "Interruption Free Zone",
                "Effective Care Transitions",
                "CARE",
                "Clear Communications",
                "Clarifying Information"
            ]
            low_content = content.lower()
            for beh in behaviors:
                if beh.lower() in low_content:
                    main_safety_behavior = beh
                    break

        if not primary_active_failure:
            print(f"Warning: Could not extract Primary Active Failure from response:\n{content}")
            primary_active_failure = "Error: Could not parse from model response"
        if not main_safety_behavior:
            print(f"Warning: Could not extract Main Safety Behavior from response:\n{content}")
            main_safety_behavior = "Error: Could not parse from model response"

        return primary_active_failure, main_safety_behavior

    except Exception as e:
        traceback.print_exc()
        return (
            f"Error identifying primary active failure: {e}",
            f"Error identifying safety behavior: {e}"
        )

def extract_text_from_pdf(pdf_path: str) -> str:
    text_content = []
    try:
        with open(pdf_path, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            for page_num in range(len(pdf_reader.pages)):
                page_text = pdf_reader.pages[page_num].extract_text()
                if page_text:
                    text_content.append(page_text)
    except Exception as e:
        traceback.print_exc()
    return "\n".join(text_content)

def parse_case_studies_from_text(pdf_text: str):
    lines = pdf_text.splitlines()
    current_title = None
    case_text      = []
    collecting     = False

    for idx, raw in enumerate(lines):
        line = raw.strip()
        low  = line.lower()

        if any(low.startswith(pref) for pref in CASE_HEADING_PREFIXES):
            # If we were previously collecting, yield the current chunk before starting a new one
            if current_title and case_text:
                yield current_title, "\n".join(case_text)

            title = "Unknown Title"
            for j in range(idx-1, -1, -1):
                prev = lines[j].strip()
                if prev:
                    title = prev.rstrip('.')
                    break

            # MODIFICATION: If the heading line itself starts with "case study" or "clinical sequence",
            # use that line as the title instead of the previous line.
            # We do this so we capture titles like "Case Study Inappropriate Anesthesia..."
            if any(x in low for x in ["case study", "clinical sequence"]):
                title = raw

            current_title = title
            case_text     = [line]
            collecting    = True
        elif collecting:
            case_text.append(line)

    if current_title and case_text:
        yield current_title, "\n".join(case_text)

def append_case_studies_to_excel(case_studies_data, excel_path=EXCEL_FILE_PATH):
    try:
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Case Studies"
            headers = [
                "Case Study #",
                "Tittle",
                "Full Case Study",
                "Summarized Case Study (200 words)",
                "Primary Active Failure",
                "Main Safety Behavior that could have prevented active failure",
                "Type Of Failure",
                "Latent Condition",
                "Casual Statement"
            ]
            ws.append(headers)
            wb.save(excel_path)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # The original code sets row_start = ws.max_row + 1
        # MODIFICATION: If the user wants these appended after row 662, ensure we don't write above row 663
        row_start = max(ws.max_row + 1, 663)  # MODIFICATION

        for i, data_item in enumerate(case_studies_data, start=row_start):
            ws.cell(row=i, column=1, value=data_item["Case Study #"])
            ws.cell(row=i, column=2, value=data_item["Tittle"])
            ws.cell(row=i, column=3, value=data_item["Full Case Study"])
            ws.cell(row=i, column=4, value=data_item["Summarized Case Study (200 words)"])
            ws.cell(row=i, column=5, value=data_item["Primary Active Failure"])
            ws.cell(row=i, column=6, value=data_item["Main Safety Behavior that could have prevented active failure"])
            ws.cell(row=i, column=7, value=data_item["Type Of Failure"])
            ws.cell(row=i, column=8, value=data_item["Latent Condition"])
            ws.cell(row=i, column=9, value=data_item["Casual Statement"])

        wb.save(excel_path)
    except Exception:
        traceback.print_exc()

def main():
    try:
        if not os.path.exists(PDF_DIRECTORY):
            print(f"ERROR: PDF directory '{PDF_DIRECTORY}' does not exist!")
            return

        def page_key(f):
            m = re.search(r'Page\s*([0-9]+)', f, re.IGNORECASE)
            return int(m.group(1)) if m else f.lower()

        pdf_files = sorted(
            (os.path.join(PDF_DIRECTORY, f) for f in os.listdir(PDF_DIRECTORY) if f.lower().endswith(".pdf")),
            key=lambda p: page_key(os.path.basename(p))
        )

        if not pdf_files:
            print(f"WARNING: No PDF files found in '{PDF_DIRECTORY}'")
            return

        merged_text = ""
        for pdf_file in pdf_files:
            print(f"Reading: {pdf_file}")
            merged_text += extract_text_from_pdf(pdf_file) + "\n"

        excel_rows_to_append = []
        case_study_counter = 2

        for title, full_case_text in parse_case_studies_from_text(merged_text):
            print(f"Processing case study: {title}")
            if not full_case_text.strip():
                print(f"WARNING: Empty case text for {title}, skipping")
                continue

            print("  Generating summary...")
            summary_200_words = summarize_case_study(full_case_text)

            print("  Identifying primary active failure and safety behavior...")
            paf, msb = identify_primary_active_failure_and_safety_behavior(full_case_text)

            row_dict = {
                "Case Study #": str(case_study_counter),
                "Tittle": title,
                "Full Case Study": full_case_text,
                "Summarized Case Study (200 words)": summary_200_words,
                "Primary Active Failure": paf,
                "Main Safety Behavior that could have prevented active failure": msb,
                "Type Of Failure": "",
                "Latent Condition": "",
                "Casual Statement": ""
            }
            excel_rows_to_append.append(row_dict)
            case_study_counter += 1

        if excel_rows_to_append:
            append_case_studies_to_excel(excel_rows_to_append, EXCEL_FILE_PATH)
            print(f"Appended {len(excel_rows_to_append)} case study entries to {EXCEL_FILE_PATH}")
        else:
            print("No case studies found to append.")

    except Exception:
        traceback.print_exc()

if __name__ == "__main__":
    main()
