from openai import OpenAI
import openpyxl
import traceback

# You can still provide your API key however you'd like:
# (either hard-coded or via environment variables).
API_KEY = "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"

# Initialize the OpenAI client using the new structure.
client = OpenAI(api_key=API_KEY)

# Path to the Excel file to update
EXCEL_FILE_PATH = "case_studies_part1_100.xlsx"

# Prompt template for classifying primary active failures
PROMPT_TEMPLATE = (
    "Classify the following primary active failure into one of three categories:"
    " Skill-Based Error, Rule-Based Error, or Knowledge-Based Error."
    " Then further classify the primary active failure from one of the subcategory from the list for the given type."
    "\n\nPrimary Active Failure: {failure}\n"
    "\nReturn the result in the exact format:\n"
    "Gem Human Error Type: <type>\n"
    "Gem Human Error Subcategory: <subcategory>"
)

# Mapping of possible subcategories for reference (not strictly needed for the prompt)
SUBCATEGORIES = {
    "Skill-Based Error": ["Slip", "Lapse", "Fumble"],
    "Rule-Based Error": [
        "Wrong Rule",
        "Misapplication of Correct Rule",
        "Non-Compliance: Low Risk Awareness",
        "Non-Compliance: Improper Coworker Coaching"
    ],
    "Knowledge-Based Error": ["Operating Outside of Expertise", "No Rule"]
}


def classify_gem_human_error(failure_text: str) -> (str, str):
    """
    Ask the OpenAI model to classify the given failure into a Gem Human Error type and subcategory.
    """
    prompt = PROMPT_TEMPLATE.format(failure=failure_text)
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=200,
            temperature=0.0
        )
        content = resp.choices[0].message.content.strip()

        # Parse the response
        error_type = None
        subcategory = None
        for line in content.splitlines():
            if line.startswith("Gem Human Error Type:"):
                error_type = line.split("Gem Human Error Type:", 1)[1].strip()
            elif line.startswith("Gem Human Error Subcategory:"):
                subcategory = line.split("Gem Human Error Subcategory:", 1)[1].strip()

        if not error_type or not subcategory:
            raise ValueError(f"Unexpected response format: {content}")

        return error_type, subcategory
    except Exception as e:
        traceback.print_exc()
        return "Error", "Error"


def main():
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    # Update headers for columns G and H
    ws.cell(row=1, column=7, value="Gem Human Error Type")
    ws.cell(row=1, column=8, value="Gem Human Error Subcategory")

    # Iterate over each primary active failure in column E starting from row 2
    for row in range(2, ws.max_row + 1):
        failure = ws.cell(row=row, column=5).value  # Column E
        if failure:
            print(f"Classifying row {row}: {failure}")
            err_type, subcat = classify_gem_human_error(failure)
            ws.cell(row=row, column=7, value=err_type)  # Column G
            ws.cell(row=row, column=8, value=subcat)   # Column H

    # Save the updated workbook
    wb.save(EXCEL_FILE_PATH)
    print("Classification complete. Workbook saved.")


if __name__ == "__main__":
    main()
